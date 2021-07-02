from flask import Flask, render_template, request
from openpyxl import Workbook
from openpyxl import load_workbook


app = Flask(__name__)

@app.route('/')
def homepage():
    excels = load_workbook('inventar.xlsx')
    page = excels["Sheet"]
    lst = []
    i = 1
    while page["A"+str(i)].value != None:
        txt = page["A" + str(i)].value 
        lst.append(txt)
        i+=1                                           

    return render_template('index.html', goods = lst)


@app.route('/add/', methods=["POST"])
def add():
    good = request.form["good"]
    excels = load_workbook('inventar.xlsx')
    page = excels["Sheet"]
    i = 1
    while page["A"+str(i)].value != None:
        i+=1 
    page["A" + str(i)].value = good
        
    excels.save('inventar.xlsx')
    return """
           <h1>Инвентарь пополнен</h1>
           <a href='/'>Домой</a>
           """

